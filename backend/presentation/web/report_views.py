from fastapi import APIRouter, Query, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
import io
import os
import sqlite3
from typing import Optional
from datetime import datetime
from openpyxl import Workbook
from fpdf import FPDF
from backend.database import sqlite_db
import logging
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from backend.presentation.middlewares.auth import login_required

router = APIRouter()
logger = logging.getLogger(__name__)

def format_vietnamese_datetime(dt_str, include_time=True):
    if not dt_str:
        return ""
    try:
        # Cắt bỏ phần millisecond nếu có (VD: 2026-05-21 15:30:00.123456)
        clean_str = str(dt_str).split('.')[0]
        dt = datetime.strptime(clean_str, "%Y-%m-%d %H:%M:%S")
        days = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
        day_name = days[dt.weekday()]
        if include_time:
            return f"{dt.strftime('%H:%M')} - {day_name} - {dt.day:02d} - {dt.month:02d} - {dt.year}"
        else:
            return f"{day_name} - {dt.day:02d} - {dt.month:02d} - {dt.year}"
    except Exception:
        return str(dt_str)

class PDFReport(FPDF):
    def __init__(self, title="Báo cáo", exporter_name="Hệ thống", date_range_str=""):
        super().__init__()
        self.report_title = title
        self.exporter_name = exporter_name
        self.date_range_str = date_range_str
        # Thêm font tiếng Việt
        font_path = os.path.join(os.getcwd(), "frontend", "static", "fonts", "Roboto-Regular.ttf")
        bold_font_path = os.path.join(os.getcwd(), "frontend", "static", "fonts", "Roboto-Bold.ttf")
        if os.path.exists(font_path):
            self.add_font("Roboto", "", font_path)
            self.add_font("Roboto", "B", bold_font_path)
            self.set_font("Roboto", "", 12)
        else:
            logger.warning(f"Không tìm thấy font tại {font_path}")
            self.set_font("Arial", "", 12)

    def header(self):
        self.set_font("Roboto", "B", 16)
        self.cell(0, 10, self.report_title, align="C", new_x="LMARGIN", new_y="NEXT")
        
        self.set_font("Roboto", "", 10)
        if self.date_range_str:
            self.cell(0, 6, self.date_range_str, align="C", new_x="LMARGIN", new_y="NEXT")
            
        info_str = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Xuất bởi: {self.exporter_name}"
        self.cell(0, 6, info_str, align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(10)

def fetch_report_data(report_type: str, start_date: str, end_date: str, camera_id: str):
    query = ""
    params = []
    
    if report_type == "traffic":
        query = """
            SELECT t.ngay_ghi_nhan as 'Ngày', c.ten_camera as 'Camera', t.so_luong_xe as 'Số lượng xe'
            FROM thong_ke_giao_thong t 
            LEFT JOIN camera c ON t.id_camera = c.id 
            WHERE 1=1
        """
        if start_date:
            query += " AND t.ngay_ghi_nhan >= ?"
            params.append(start_date)
        if end_date:
            query += " AND t.ngay_ghi_nhan <= ?"
            params.append(end_date)
        if camera_id and camera_id != "all":
            query += " AND t.id_camera = ?"
            params.append(int(camera_id))
        query += " ORDER BY t.ngay_ghi_nhan DESC"
            
    elif report_type == "parking":
        query = """
            SELECT p.thoi_gian_vi_pham as 'Bắt đầu', IFNULL(p.thoi_gian_ket_thuc, 'Đang diễn ra') as 'Kết thúc',
                   c.ten_camera as 'Camera', 
                   IFNULL(c.mo_ta, '') as 'Địa điểm',
                   IFNULL(p.bien_so, 'Không xác định') as 'Biển số', 
                   CASE WHEN p.da_giai_quyet = 1 THEN 'Đã xử lý' ELSE 'Chưa xử lý' END as 'Trạng thái',
                   p.duong_dan_anh as 'Ảnh'
            FROM vi_pham_do_xe p 
            LEFT JOIN camera c ON p.id_camera = c.id 
            WHERE 1=1
        """
        if start_date:
            query += " AND DATE(p.thoi_gian_vi_pham) >= ?"
            params.append(start_date)
        if end_date:
            query += " AND DATE(p.thoi_gian_vi_pham) <= ?"
            params.append(end_date)
        if camera_id and camera_id != "all":
            query += " AND p.id_camera = ?"
            params.append(int(camera_id))
        query += " ORDER BY c.ten_camera, p.thoi_gian_vi_pham DESC"

    elif report_type == "congestion":
        query = """
            SELECT n.thoi_gian_bat_dau as 'Bắt đầu', IFNULL(n.thoi_gian_ket_thuc, 'Đang diễn ra') as 'Kết thúc', 
                   c.ten_camera as 'Camera', IFNULL(c.mo_ta, '') as 'Địa điểm', n.muc_do_un_tac as 'Mức độ', 
                   IFNULL(n.thoi_gian_keo_dai_giay, 0) as 'Kéo dài (giây)',
                   n.duong_dan_anh as 'Ảnh'
            FROM nhat_ky_un_tac n 
            LEFT JOIN camera c ON n.id_camera = c.id 
            WHERE 1=1
        """
        if start_date:
            query += " AND DATE(n.thoi_gian_bat_dau) >= ?"
            params.append(start_date)
        if end_date:
            query += " AND DATE(n.thoi_gian_bat_dau) <= ?"
            params.append(end_date)
        if camera_id and camera_id != "all":
            query += " AND n.id_camera = ?"
            params.append(int(camera_id))
        query += " ORDER BY n.thoi_gian_bat_dau ASC"
    else:
        raise ValueError("Invalid report type")

    with sqlite_db.connect() as conn:
        cursor = conn.execute(query, params)
        columns = [description[0] for description in cursor.description]
        rows = cursor.fetchall()
        
    formatted_data = []
    
    # Process data and STT for traffic
    if report_type == "traffic":
        columns.insert(0, "STT")
        total_xe = 0
        for idx, row in enumerate(rows):
            d = dict(row)
            d["STT"] = idx + 1
            if d.get("Ngày"):
                d["Ngày"] = format_vietnamese_datetime(d["Ngày"], include_time=False)
            total_xe += int(d.get("Số lượng xe", 0))
            formatted_data.append(d)
        
        # Add total row
        total_row = {col: "" for col in columns}
        total_row["STT"] = ""
        total_row["Ngày"] = "Tổng cộng"
        total_row["Số lượng xe"] = total_xe
        formatted_data.append(total_row)
        

    else:
        if "Bắt đầu" in columns:
            idx = columns.index("Bắt đầu")
            columns.insert(idx, "Thời gian")
            columns.remove("Bắt đầu")
            if "Kết thúc" in columns: 
                columns.remove("Kết thúc")
                
        for row in rows:
            d = dict(row)
            start_str = str(d.get("Bắt đầu", ""))
            end_str = str(d.get("Kết thúc", "Đang diễn ra"))
            try:
                start_dt = datetime.strptime(start_str.split('.')[0], "%Y-%m-%d %H:%M:%S")
                start_time = start_dt.strftime('%H:%M:%S')
                days = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                day_name = days[start_dt.weekday()]
                date_str = f"{day_name} {start_dt.day:02d}-{start_dt.month:02d}-{start_dt.year}"
                
                if end_str and end_str != "Đang diễn ra":
                    end_dt = datetime.strptime(end_str.split('.')[0], "%Y-%m-%d %H:%M:%S")
                    end_time = end_dt.strftime('%H:%M:%S')
                    d["Thời gian"] = f"Từ {start_time} -> {end_time}, {date_str}"
                else:
                    d["Thời gian"] = f"Từ {start_time} -> Đang diễn ra, {date_str}"
            except Exception:
                d["Thời gian"] = f"Từ {start_str} -> {end_str}"
                
            if "Bắt đầu" in d: del d["Bắt đầu"]
            if "Kết thúc" in d: del d["Kết thúc"]
            formatted_data.append(d)
            
    return columns, formatted_data

from openpyxl.chart import BarChart, Reference
import uuid

def create_excel(columns, data, report_title, report_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo"
    
    # Remove 'Ảnh' from columns for Excel if present
    cols_to_write = [c for c in columns if c != 'Ảnh']
    
    # Tiêu đề cột
    ws.append(cols_to_write)
    
    # Dữ liệu
    for row in data:
        ws.append([row[col] for col in cols_to_write])
        
    if report_type == "traffic" and len(data) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Biểu đồ lưu lượng giao thông"
        chart.y_axis.title = 'Số lượng xe'
        chart.x_axis.title = 'Ngày'
        
        # Reference cho dữ liệu số lượng xe (cột cuối cùng)
        # Bỏ qua dòng "Tổng cộng" ở cuối cùng
        data_ref = Reference(ws, min_col=len(cols_to_write), min_row=1, max_row=len(data)) 
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(data)) # Cột Ngày
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        ws.add_chart(chart, f"A{len(data) + 3}")
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_pdf(columns, data, report_title, report_type, exporter_name, date_range_str, extra_data=None):
    pdf = PDFReport(title=report_title, exporter_name=exporter_name, date_range_str=date_range_str)
    pdf.add_page()
    
    if report_type == "parking":
        # ═══════════════════════════════════════════════════════════════
        # BÁO CÁO VI PHẠM DỪNG ĐỖ XE — Mỗi vi phạm 1 trang, gom nhóm theo Camera
        # ═══════════════════════════════════════════════════════════════
        
        # --- TRANG TỔNG QUAN (SUMMARY) ---
        total_violations = len(data)
        cameras_set = set()
        resolved_count = 0
        for row in data:
            cam_name = row.get('Camera', 'N/A')
            if cam_name:
                cameras_set.add(cam_name)
            if row.get('Trạng thái', '') == 'Đã xử lý':
                resolved_count += 1
        unresolved_count = total_violations - resolved_count
        
        pdf.set_font("Roboto", "B", 14)
        pdf.cell(0, 10, "TỔNG QUAN", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        
        summary_data = [
            ["Tổng số vi phạm", str(total_violations)],
            ["Số camera phát hiện", str(len(cameras_set))],
            ["Đã xử lý", str(resolved_count)],
            ["Chưa xử lý", str(unresolved_count)],
        ]
        try:
            with pdf.table(col_widths=[120, 70], text_align="CENTER", line_height=10) as table:
                header = table.row()
                header.cell("Thông tin")
                header.cell("Giá trị")
                for item in summary_data:
                    row_t = table.row()
                    row_t.cell(item[0])
                    row_t.cell(item[1])
        except AttributeError:
            pdf.set_font("Roboto", "", 11)
            for item in summary_data:
                pdf.cell(120, 10, item[0], border=1)
                pdf.cell(70, 10, item[1], border=1, align="C")
                pdf.ln()
        
        pdf.ln(10)
        
        # --- GOM NHÓM THEO CAMERA ---
        grouped = {}
        for row in data:
            cam_name = row.get('Camera', 'Không xác định')
            if cam_name not in grouped:
                grouped[cam_name] = {"location": row.get('Địa điểm', ''), "violations": []}
            grouped[cam_name]["violations"].append(row)
        
        violation_counter = 0
        for cam_name, cam_data in grouped.items():
            location = cam_data["location"]
            violations = cam_data["violations"]
            
            for vi in violations:
                violation_counter += 1
                pdf.add_page()
                
                # Tiêu đề Camera
                pdf.set_font("Roboto", "B", 14)
                pdf.set_x(10)
                pdf.multi_cell(w=190, h=8, txt=f"Camera: {cam_name}")
                if location:
                    pdf.set_font("Roboto", "", 10)
                    pdf.set_x(10)
                    pdf.multi_cell(w=190, h=6, txt=f"Địa điểm: {location}")
                
                # Đường kẻ phân cách
                pdf.ln(3)
                pdf.set_draw_color(100, 100, 100)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(5)
                
                # Thông tin vi phạm
                pdf.set_font("Roboto", "B", 13)
                pdf.set_x(10)
                bien_so = vi.get('Biển số', 'N/A')
                pdf.multi_cell(w=190, h=9, txt=f"Vi phạm #{violation_counter} — Biển số: {bien_so}")
                pdf.ln(3)
                
                pdf.set_font("Roboto", "", 11)
                pdf.set_x(10)
                pdf.multi_cell(w=190, h=7, txt=f"Thời gian: {vi.get('Thời gian', 'N/A')}")
                pdf.set_x(10)
                pdf.multi_cell(w=190, h=7, txt=f"Trạng thái: {vi.get('Trạng thái', 'N/A')}")
                
                # Ảnh bằng chứng
                img_path = vi.get('Ảnh')
                if img_path:
                    full_path = os.path.join(os.getcwd(), img_path.strip('/'))
                    if os.path.exists(full_path):
                        pdf.ln(5)
                        # Tính toán kích thước ảnh nhỏ lại để không tràn trang (rộng 120mm, căn giữa: (210-120)/2 = 45)
                        pdf.image(full_path, x=45, w=120)
                    else:
                        pdf.ln(3)
                        pdf.set_x(10)
                        pdf.multi_cell(w=190, h=7, txt="Hình ảnh: Không có (Lỗi file)")
                else:
                    pdf.ln(3)
                    pdf.set_x(10)
                    pdf.multi_cell(w=190, h=7, txt="Hình ảnh: Không có")
    
    elif report_type == "congestion":
        # ═══════════════════════════════════════════════════════════════
        # BÁO CÁO ÙN TẮC — Danh sách theo thời gian
        # ═══════════════════════════════════════════════════════════════
        for idx, row in enumerate(data):
            pdf.set_font("Roboto", "B", 13)
            pdf.set_x(10)
            pdf.multi_cell(w=190, h=8, txt=f"Bản ghi #{idx+1} — Mức độ: {row.get('Mức độ', 'N/A')}")
            
            pdf.set_font("Roboto", "", 11)
            pdf.set_x(10)
            pdf.multi_cell(w=190, h=7, txt=f"Camera: {row.get('Camera', 'N/A')}")
            pdf.set_x(10)
            pdf.multi_cell(w=190, h=7, txt=f"Địa điểm: {row.get('Địa điểm', 'N/A')}")
            pdf.set_x(10)
            pdf.multi_cell(w=190, h=7, txt=f"Thời gian: {row.get('Thời gian', 'N/A')}")
            pdf.set_x(10)
            pdf.multi_cell(w=190, h=7, txt=f"Kéo dài: {row.get('Kéo dài (giây)', 0)} giây")
            
            img_path = row.get('Ảnh')
            if img_path:
                full_path = os.path.join(os.getcwd(), img_path.strip('/'))
                if os.path.exists(full_path):
                    pdf.ln(3)
                    pdf.image(full_path, x=45, w=120)
                    pdf.ln(5)
                else:
                    pdf.set_x(10)
                    pdf.multi_cell(w=190, h=7, txt="Hình ảnh: Không có (Lỗi file)")
            else:
                pdf.set_x(10)
                pdf.multi_cell(w=190, h=7, txt="Hình ảnh: Không có")
                
            pdf.ln(8)
            pdf.set_draw_color(200, 200, 200)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(5)
    else:
        # ═══════════════════════════════════════════════════════════════
        # BÁO CÁO LƯU LƯỢNG GIAO THÔNG — Chia bảng theo Camera
        # ═══════════════════════════════════════════════════════════════
        
        # Ánh xạ tên loại xe sang tiếng Việt
        type_map = {'car': 'Ô tô', 'truck': 'Xe tải', 'bus': 'Xe khách', 'motorcycle': 'Xe máy', 'bicycle': 'Xe đạp'}
        
        # Biểu đồ tổng hợp (giữ nguyên logic cũ)
        cols_to_write = [c for c in columns if c != 'Ảnh']
        if len(data) > 1:
            try:
                daily_counts = {}
                for row in data[:-1]:
                    parts = str(row['Ngày']).split(' ')
                    if len(parts) >= 7:
                        date_str = f"{parts[2]}/{parts[4]}/{parts[6]}"
                    else:
                        date_str = str(row['Ngày'])[:10]
                        
                    daily_counts[date_str] = daily_counts.get(date_str, 0) + int(row['Số lượng xe'])
                
                dates = list(daily_counts.keys())
                counts = list(daily_counts.values())
                
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5), gridspec_kw={'width_ratios': [2, 1.2]})
                
                bars = ax1.bar(dates, counts, color='#3b82f6')
                ax1.bar_label(bars, fmt='%d', padding=3, rotation=90, fontsize=8)
                if counts:
                    ax1.set_ylim(0, max(counts) * 1.3)
                    
                ax1.set_title('Biểu đồ lưu lượng giao thông', fontsize=12)
                ax1.set_xlabel('Ngày')
                ax1.set_ylabel('Số lượng xe')
                ax1.tick_params(axis='x', rotation=45, labelsize=9)
                
                if extra_data and "vehicle_dist" in extra_data and extra_data["vehicle_dist"]:
                    dist = extra_data["vehicle_dist"]
                    labels = [f"{type_map.get(d['type'], d['type'])} ({d['count']})" for d in dist]
                    sizes = [d['count'] for d in dist]
                    colors = ['#2563eb', '#06b6d4', '#10b981', '#f59e0b', '#f43f5e']
                    
                    wedges, texts, autotexts = ax2.pie(sizes, autopct=lambda p: f'{p:.1f}%' if p > 3 else '', startangle=90, colors=colors[:len(labels)], pctdistance=0.85)
                    ax2.legend(wedges, labels, title="Loại xe", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
                    ax2.set_title('Tỉ lệ loại xe', fontsize=12)
                else:
                    ax2.axis('off')
                
                plt.tight_layout()
                chart_filename = f"temp_chart_{uuid.uuid4().hex}.png"
                plt.savefig(chart_filename)
                plt.close()
                
                pdf.image(chart_filename, x=5, w=190)
                pdf.ln(5)
                os.remove(chart_filename)
            except Exception as e:
                logger.error(f"Lỗi vẽ biểu đồ: {e}")
        
        # --- BẢNG CHI TIẾT CHIA THEO CAMERA ---
        traffic_by_camera = extra_data.get("traffic_by_camera", []) if extra_data else []
        
        if traffic_by_camera:
            # Gom nhóm dữ liệu theo camera
            camera_groups = {}
            for row in traffic_by_camera:
                cam_id = row["camera_id"]
                if cam_id not in camera_groups:
                    camera_groups[cam_id] = {
                        "name": row["camera_name"] or f"Camera #{cam_id}",
                        "description": row["description"],
                        "vehicles": []
                    }
                camera_groups[cam_id]["vehicles"].append({
                    "type": type_map.get(row["vehicle_type"], row["vehicle_type"]),
                    "count": row["count"]
                })
            
            for cam_id, cam_data in camera_groups.items():
                pdf.ln(8)
                
                # Tiêu đề camera
                pdf.set_font("Roboto", "B", 13)
                pdf.set_x(10)
                pdf.multi_cell(w=190, h=8, txt=f"Camera: {cam_data['name']}")
                if cam_data["description"]:
                    pdf.set_font("Roboto", "", 10)
                    pdf.set_x(10)
                    pdf.multi_cell(w=190, h=6, txt=f"Địa điểm: {cam_data['description']}")
                pdf.ln(3)
                
                # Bảng loại xe
                total_count = sum(v["count"] for v in cam_data["vehicles"])
                try:
                    with pdf.table(col_widths=[20, 100, 70], text_align="CENTER", line_height=8) as table:
                        header = table.row()
                        header.cell("STT")
                        header.cell("Loại xe")
                        header.cell("Số lượng")
                        
                        for idx, v in enumerate(cam_data["vehicles"], start=1):
                            row_t = table.row()
                            row_t.cell(str(idx))
                            row_t.cell(v["type"])
                            row_t.cell(str(v["count"]))
                        
                        # Dòng tổng cộng
                        total_row = table.row()
                        total_row.cell("")
                        total_row.cell("Tổng cộng")
                        total_row.cell(str(total_count))
                except AttributeError:
                    # Fallback cho fpdf2 cũ
                    pdf.set_font("Roboto", "B", 10)
                    for col_name, col_w in [("STT", 20), ("Loại xe", 100), ("Số lượng", 70)]:
                        pdf.cell(col_w, 10, col_name, border=1, align="C")
                    pdf.ln()
                    pdf.set_font("Roboto", "", 10)
                    for idx, v in enumerate(cam_data["vehicles"], start=1):
                        pdf.cell(20, 10, str(idx), border=1, align="C")
                        pdf.cell(100, 10, v["type"], border=1, align="C")
                        pdf.cell(70, 10, str(v["count"]), border=1, align="C")
                        pdf.ln()
                    pdf.set_font("Roboto", "B", 10)
                    pdf.cell(20, 10, "", border=1)
                    pdf.cell(100, 10, "Tổng cộng", border=1, align="C")
                    pdf.cell(70, 10, str(total_count), border=1, align="C")
                    pdf.ln()
        else:
            # Fallback: Nếu không có dữ liệu chi tiết theo camera, hiển thị bảng cũ
            pdf.set_font("Roboto", "", 10)
            col_widths = [190 / len(cols_to_write)] * len(cols_to_write)
            col_widths[0] = 15
            col_widths[1] = 75
            col_widths[2] = 70
            col_widths[3] = 30
                
            try:
                with pdf.table(text_align="CENTER", col_widths=col_widths, line_height=8) as table:
                    row = table.row()
                    for col in cols_to_write:
                        row.cell(str(col))
                    for data_row in data:
                        row = table.row()
                        for col in cols_to_write:
                            row.cell(str(data_row[col]))
            except AttributeError:
                pdf.set_font("Roboto", "B", 10)
                for i, col in enumerate(cols_to_write):
                    pdf.cell(col_widths[i], 10, str(col), border=1, align="C")
                pdf.ln()
                pdf.set_font("Roboto", "", 10)
                for r in data:
                    for i, col in enumerate(cols_to_write):
                        val = str(r[col])
                        pdf.cell(col_widths[i], 10, val[:40], border=1, align="C")
                    pdf.ln()
            
    output = io.BytesIO(pdf.output(dest="S"))
    output.seek(0)
    return output

@router.post("/export")
@router.get("/export")
async def export_report(
    request: Request,
    report_type: str = Query(..., description="traffic, parking, congestion"),
    format: str = Query(..., description="excel or pdf"),
    start_date: str = Query(""),
    end_date: str = Query(""),
    camera_id: str = Query("all"),
    user = Depends(login_required)
):
    titles = {
        "traffic": "BÁO CÁO LƯU LƯỢNG GIAO THÔNG",
        "parking": "BÁO CÁO VI PHẠM DỪNG ĐỖ XE TRÁI QUY ĐỊNH",
        "congestion": "BÁO CÁO TÌNH HÌNH ÙN TẮC GIAO THÔNG"
    }
    
    if report_type not in titles:
        raise HTTPException(status_code=400, detail="Loại báo cáo không hợp lệ")
        
    try:
        columns, data = fetch_report_data(report_type, start_date, end_date, camera_id)
        report_title = titles[report_type]
        
        # User name and Date range
        exporter_name = getattr(user, "full_name", getattr(user, "username", "Hệ thống"))
        if not exporter_name:
            exporter_name = getattr(user, "username", "Hệ thống")
            
        date_range_str = ""
        if start_date and end_date:
            date_range_str = f"Từ ngày {start_date} đến ngày {end_date}"
        elif start_date:
            date_range_str = f"Từ ngày {start_date}"
        elif end_date:
            date_range_str = f"Đến ngày {end_date}"
        
        filename_prefix = f"report_{report_type}_{datetime.now().strftime('%Y%m%d%H%M')}"
        
        extra_data = None
        if report_type == "traffic":
            from backend.database.sqlite_db import get_vehicle_type_distribution, get_traffic_report_by_camera
            cam_ids = [int(camera_id)] if camera_id and camera_id != "all" else None
            extra_data = {
                "vehicle_dist": get_vehicle_type_distribution(start_date, end_date, cam_ids),
                "traffic_by_camera": get_traffic_report_by_camera(start_date, end_date, cam_ids),
            }
        
        if format == "excel":
            file_stream = create_excel(columns, data, report_title, report_type)
            return StreamingResponse(
                file_stream, 
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename_prefix}.xlsx"}
            )
        elif format == "pdf":
            file_stream = create_pdf(columns, data, report_title, report_type, exporter_name, date_range_str, extra_data)
            return StreamingResponse(
                file_stream, 
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename_prefix}.pdf"}
            )
        else:
            raise HTTPException(status_code=400, detail="Định dạng không hợp lệ")
            
    except Exception as e:
        logger.error(f"Lỗi xuất báo cáo: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi hệ thống khi xuất báo cáo: {str(e)}")
